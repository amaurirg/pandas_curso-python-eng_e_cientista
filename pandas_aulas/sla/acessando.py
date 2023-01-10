import openpyxl
from copy import copy
import os

# filename_xls = 'rel_sla_teste.xls'
# os.system(f'libreoffice --convert-to xlsx {filename_xls} --headless')

filename = "rel_sla_teste.xlsx"
wb = openpyxl.load_workbook(filename)
sheet = wb['Sheet1']
ws = wb.active

ws.column_dimensions['V'].hidden = False
ws.unmerge_cells('P11:U11')
ws.merge_cells('P11:V11')

sheet['V12'].alignment = copy(sheet['T12'].alignment)
sheet['V12'].font = copy(sheet['T12'].font)
# sheet['V11'].border = copy(sheet['A11'].border)
sheet['V12'].border = copy(sheet['T12'].border)
sheet['V12'] = 'Destino'

for linha in sheet['V13':'V23']:
    for celula in linha:
        celula.value = 'Nacional'
        celula.alignment = copy(sheet['A14'].alignment)
        celula.font = copy(sheet['A14'].font)

for linha in sheet['V26':'V35']:
    for celula in linha:
        celula.value = 'Internacional'
        celula.alignment = copy(sheet['A14'].alignment)
        celula.font = copy(sheet['A14'].font)

ws.unmerge_cells('A13:U13')
ws.unmerge_cells('A25:U25')
ws.delete_rows(24, 2)
ws.delete_rows(13, 1)

wb.save(filename)
