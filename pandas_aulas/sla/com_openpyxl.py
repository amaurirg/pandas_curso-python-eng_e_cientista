#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.
# ! python3
# Atualizando valores em planilhas xlsx

import openpyxl  # importa o módulo openpyxl para trabalhar com planilhas no formato xlsx

wb = openpyxl.load_workbook('preparando_arquivo/rel_sla_teste.xls')  # lê o documento, no caso, produceSales.xlsx
sheet = wb.get_sheet_by_name('Sheet1')  # informa para trabalhar na planilha especificada, no caso, Sheet

# for rowNum in range(2, sheet.max_row + 1):  # percorre o documento da 2ª linha até a última. Tem que colocar '+1' pois o range seria até a penúltima linha
#     produceName = sheet.cell(row=rowNum,
#                              column=1).value  # percorre as linhas da coluna 1 e armazena o valor da célula em 'produceName'
#     if produceName in PRICE_UPDATES:  # se o produto for encontrado na coluna 1 executa o comando abaixo
#         sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]  # altera o valor do produto na coluna 2

wb.save('updated_rel_sla_teste.xlsx')  # comando necessário para salvar efetivamente as alterações na planilha