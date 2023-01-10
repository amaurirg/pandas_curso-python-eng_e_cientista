import openpyxl
from copy import copy


filename = "rel_sla_teste.xlsx"
wb = openpyxl.load_workbook(filename)
sheet = wb['Sheet1']
ws = wb.active

ws.column_dimensions['V'].hidden = False
ws.unmerge_cells('P11:U11')
ws.merge_cells('P11:V11')

sheet['V12'].alignment = copy(sheet['T12'].alignment)
sheet['V12'].font = copy(sheet['T12'].font)
sheet['V12'].border = copy(sheet['T12'].border)
sheet['V12'] = 'Destino'

# sheet['V11'].border = copy(sheet['A11'].border)
rows_nac = {}
rows_internac = {}

for cell_item in ws['A']:
    if cell_item.value == 'Nacional':
        rows_nac['cell'] = cell_item.row
        rows_nac['start'] = cell_item.row + 1
    elif cell_item.value == 'Internacional':
        rows_internac['cell'] = cell_item.row
        rows_internac['start'] = cell_item.row + 1
    elif cell_item.value is None:
        if len(rows_nac) == 2:
            rows_nac['end'] = cell_item.row - 1
        elif len(rows_internac) == 2:
            rows_internac['end'] = cell_item.row - 1

for line in sheet[f'V{rows_nac["start"]}': f'V{rows_nac["end"]}']:
    for cel_nac in line:
        cel_nac.value = 'Nacional'
        cel_nac.alignment = copy(sheet[f'A{rows_nac["start"]}'].alignment)
        cel_nac.font = copy(sheet[f'A{rows_nac["start"]}'].font)

for line in sheet[f'V{rows_internac["start"]}': f'V{rows_internac["end"]}']:
    for cell_internac in line:
        cell_internac.value = 'Internacional'
        cell_internac.alignment = copy(sheet[f'A{rows_nac["start"]}'].alignment)
        cell_internac.font = copy(sheet[f'A{rows_nac["start"]}'].font)

ws.delete_rows(rows_internac['end'] + 1, 10)
null_lines_diff = rows_internac['cell'] - (rows_nac['end'] + 1)
ws.unmerge_cells(f'A{rows_internac["cell"]}:U{rows_internac["cell"]}')
ws.delete_rows(rows_internac['cell'])
ws.delete_rows(rows_nac['end'] + 1, 1)
ws.unmerge_cells(f'A{rows_nac["cell"]}:U{rows_nac["cell"]}')
ws.delete_rows(rows_nac['cell'])

wb.save(filename)
